"""
AI Customer Support System - Main Application
Complete solution for customer query resolution and ticket management
"""

import streamlit as st
import pandas as pd
import json
import base64
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import os
import uuid
from typing import List, Dict
import re
from openpyxl import Workbook, load_workbook
import time
from io import BytesIO
import hashlib
import secrets

# Import our modules
from categorizer import TicketCategorizer
from resolver import QueryResolver
from sheets_client import GoogleSheetsClient
from tagger import TicketTagger
from chatbot import Chatbot
from notifier import Notifier, SlackConfig, EmailConfig
from rag_engine import create_documents_from_knowledge_base

# Page configuration
st.set_page_config(
    page_title="AI Powered Knowledge Engine",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        text-align: center;
        margin-bottom: 2rem;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    .ticket-card {
        background: #2d3748;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #667eea;
        margin: 0.5rem 0;
    }
    
    .success-message {
        background: #2d5a27;
        color: #68d391;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #68d391;
    }
    
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin: 0.5rem 0;
    }
    .sub-header { text-align: center; font-size: 1.25rem; margin-top: -0.25rem; margin-bottom: 1rem; }
    /* Tabs visibility */
    .stTabs [role="tablist"] { gap: 10px; border-bottom: 1px solid rgba(255,255,255,0.15); padding: 4px 6px; }
    .stTabs [role="tab"] { padding: 6px 12px; border-radius: 8px 8px 0 0; color: #e5e7eb; background: rgba(255,255,255,0.04); font-weight: 600; }
    .stTabs [role="tab"][aria-selected="true"] { color: #ffffff; background: rgba(99,102,241,0.18); border-bottom: 3px solid #ef4444; }
</style>
""", unsafe_allow_html=True)

def compute_sentiment_label(text: str) -> str:
    """Return a coarse sentiment label for the given text.

    Heuristic approach without external dependencies:
    counts positive/negative cue words and returns Positive/Negative/Neutral.
    """
    try:
        if not isinstance(text, str) or not text.strip():
            return "Neutral"
        lowered = text.lower()
        positive_words = {
            "thanks","thank you","great","good","awesome","excellent","fixed","resolved",
            "success","working","satisfied","love","perfect","amazing","helpful"
        }
        negative_words = {
            "bad","terrible","broken","error","issue","problem","not working","can't",
            "cannot","fail","failed","fails","crash","crashes","slow","worst","angry",
            "upset","frustrated","hate","delay","delayed","refund"
        }
        pos = sum(1 for w in positive_words if w in lowered)
        neg = sum(1 for w in negative_words if w in lowered)
        if neg > pos and neg >= 1:
            return "Negative"
        if pos > neg and pos >= 1:
            return "Positive"
        return "Neutral"
    except Exception:
        return "Neutral"

def initialize_session_state():
    """Initialize session state variables."""
    if 'tickets' not in st.session_state:
        st.session_state.tickets = []
    if 'google_sheet_data' not in st.session_state:
        st.session_state.google_sheet_data = []
    if 'categorizer' not in st.session_state:
        st.session_state.categorizer = TicketCategorizer()
    if 'resolver' not in st.session_state:
        st.session_state.resolver = QueryResolver()
    if 'tagger' not in st.session_state:
        st.session_state.tagger = TicketTagger()
    if 'sheets_client' not in st.session_state:
        st.session_state.sheets_client = GoogleSheetsClient()
    if 'csv_loaded' not in st.session_state:
        st.session_state.csv_loaded = False
    if 'excel_autosave' not in st.session_state:
        st.session_state.excel_autosave = True
    if 'excel_path' not in st.session_state:
        st.session_state.excel_path = 'tickets.xlsx'
    if 'pending_query' not in st.session_state:
        st.session_state.pending_query = None
    if 'last_created_ticket_id' not in st.session_state:
        st.session_state.last_created_ticket_id = None
    if 'last_created_ticket_ai' not in st.session_state:
        st.session_state.last_created_ticket_ai = []
    if 'flash_ticket_id' not in st.session_state:
        st.session_state.flash_ticket_id = None
    if 'flash_ticket_status' not in st.session_state:
        st.session_state.flash_ticket_status = None
    if 'chatbot' not in st.session_state:
        st.session_state.chatbot = None
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []
    if 'chatbot_initialized' not in st.session_state:
        st.session_state.chatbot_initialized = False
    if 'google_api_key' not in st.session_state:
        st.session_state.google_api_key = None
    if 'google_cse_id' not in st.session_state:
        st.session_state.google_cse_id = None
    if 'notifier' not in st.session_state:
        st.session_state.notifier = Notifier(enabled=False)

# ---- Simple user store (local JSON with salted hashes) ----
USERS_FILE = 'users.json'

def _load_users() -> Dict[str, Dict[str, str]]:
    try:
        if os.path.exists(USERS_FILE):
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return {}

def _save_users(users: Dict[str, Dict[str, str]]) -> bool:
    try:
        with open(USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(users, f, indent=2)
        return True
    except Exception:
        return False

def _hash_password(password: str, salt: str) -> str:
    return hashlib.sha256((salt + password).encode('utf-8')).hexdigest()

def create_user(username: str, password: str) -> tuple[bool, str]:
    username = (username or '').strip()
    if not username or not password:
        return False, 'Username and password are required.'
    users = _load_users()
    if username in users:
        return False, 'Username already exists.'
    salt = secrets.token_hex(16)
    pwd_hash = _hash_password(password, salt)
    users[username] = { 'salt': salt, 'password_hash': pwd_hash }
    ok = _save_users(users)
    if not ok:
        return False, 'Failed to save user.'
    return True, 'User registered successfully.'

def verify_user(username: str, password: str) -> bool:
    users = _load_users()
    info = users.get((username or '').strip())
    if not info:
        return False
    salt = info.get('salt', '')
    expected = info.get('password_hash', '')
    return expected and (_hash_password(password or '', salt) == expected)

def any_registered_users() -> bool:
    return bool(_load_users())

def load_csv_data():
    """Load sample tickets from CSV file."""
    try:
        if os.path.exists('sample_tickets.csv'):
            df = pd.read_csv('sample_tickets.csv')
            
            # Convert DataFrame to list of dictionaries
            tickets = []
            for _, row in df.iterrows():
                # Parse AI response and tags from string format
                ai_response = eval(row['ai_response']) if isinstance(row['ai_response'], str) else row['ai_response']
                tags = eval(row['tags']) if isinstance(row['tags'], str) else row['tags']
                # Robust boolean parse for 'solved'
                raw_solved = row.get('solved', False)
                if isinstance(raw_solved, str):
                    solved_val = raw_solved.strip().lower() in ("true", "1", "yes", "y")
                else:
                    solved_val = bool(raw_solved)
                
                # Compute sentiment from detailed_issue if available, else from summary
                detail_text = row.get('detailed_issue')
                summary_text = row.get('issue_summary')
                sentiment_label = compute_sentiment_label(str(detail_text or summary_text or ""))

                ticket = {
                    "ticket_id": row['ticket_id'],
                    "customer_email": row['customer_email'],
                    "customer_name": row['customer_name'],
                    "issue_summary": row['issue_summary'],
                    "detailed_issue": row['detailed_issue'],
                    "category": row['category'],
                    "priority": row['priority'],
                    "status": row['status'],
                    "created_date": row['created_date'],
                    "created_time": "12:00:00",  # Default time since CSV doesn't have time
                    "platform": row['platform'],
                    "contact_type": "CSV Import",
                    "ai_response": ai_response,
                    "confidence": float(row['ai_confidence']),
                    "solved": solved_val,
                    "tags": tags,
                    "sentiment": sentiment_label
                }
                tickets.append(ticket)
            
            return tickets
        else:
            st.error("CSV file 'sample_tickets.csv' not found!")
            return []
    except Exception as e:
        st.error(f"Error loading CSV data: {str(e)}")
        return []

def load_excel_data(excel_path: str) -> List[Dict]:
    """Load tickets from an Excel file into the in-memory ticket format."""
    try:
        if not os.path.exists(excel_path):
            st.error(f"Excel file not found: {excel_path}")
            return []

        df = pd.read_excel(excel_path, engine='openpyxl')
        # Build a case-insensitive column map and allow common alternative headers
        colmap = {str(c).strip().lower(): c for c in df.columns}
        def get_val(row, *keys):
            # keys are lower-case desired names or common alternates
            for k in keys:
                kk = str(k).strip().lower()
                if kk in colmap:
                    return row.get(colmap[kk])
            return None

        tickets: List[Dict] = []
        for _, row in df.iterrows():
            # Parse JSON list fields if saved as strings
            ai_response = get_val(row, 'ai_response', 'ai response')
            if isinstance(ai_response, str):
                try:
                    ai_response = json.loads(ai_response)
                except Exception:
                    # Also support semicolon-separated strings
                    ai_response = [s.strip() for s in ai_response.split(';') if s.strip()]

            tags = get_val(row, 'tags')
            if isinstance(tags, str):
                try:
                    tags = json.loads(tags)
                except Exception:
                    tags = [t.strip() for t in tags.split(',') if t.strip()]

            ticket = {
                "ticket_id": get_val(row, 'ticket_id', 'ticket id') or f"TK-{uuid.uuid4().hex[:12].upper()}",
                "customer_email": get_val(row, 'customer_email', 'customer email') or '',
                "customer_name": get_val(row, 'customer_name', 'customer name') or '',
                "issue_summary": get_val(row, 'issue_summary', 'issue summary') or '',
                "detailed_issue": get_val(row, 'detailed_issue', 'detailed issue') or '',
                "category": get_val(row, 'category') or 'General',
                "priority": get_val(row, 'priority') or 'Medium',
                "tags": tags if isinstance(tags, list) else [],
                "status": get_val(row, 'status') or 'Open',
                "created_date": str(get_val(row, 'created_date', 'created date') or datetime.now().strftime('%Y-%m-%d')).split(' ')[0],
                "created_time": str(get_val(row, 'created_time', 'created time') or datetime.now().strftime('%H:%M:%S')),
                "ai_response": ai_response if isinstance(ai_response, list) else [],
                "confidence": float((get_val(row, 'confidence', 'ai confidence') or 0.8)),
                "solved": bool(str(get_val(row, 'solved') or '').strip().lower() in {"true","1","yes","y"}),
                "platform": get_val(row, 'platform') or 'Web',
                "contact_type": get_val(row, 'contact_type', 'contact type') or 'Web Form',
                "sentiment": get_val(row, 'sentiment') or compute_sentiment_label(str(get_val(row, 'detailed_issue', 'detailed issue') or get_val(row, 'issue_summary', 'issue summary') or ''))
            }
            tickets.append(ticket)

        return tickets
    except Exception as e:
        st.error(f"Error loading Excel data: {str(e)}")
        return []

def save_ticket_to_excel(ticket: Dict, excel_path: str) -> bool:
    """Append a single ticket to an Excel file, creating it if missing.

    Returns True on success, False otherwise.
    """
    try:
        # Open in append mode using openpyxl so we don't rewrite the file
        columns = [
            'ticket_id', 'customer_email', 'customer_name', 'issue_summary', 'detailed_issue',
            'category', 'priority', 'tags', 'status', 'created_date', 'created_time',
            'ai_response', 'confidence', 'solved', 'platform', 'contact_type', 'sentiment'
        ]
        # Ensure workbook exists and has header; migrate missing columns (e.g., 'sentiment')
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            # Initialize header if file is essentially empty
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
                ws.append(columns)
            else:
                # Read existing header row
                existing_headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
                if not existing_headers:
                    ws.append(columns)
                else:
                    # Append any missing columns to the end, and backfill blanks for prior rows
                    missing = [c for c in columns if c not in existing_headers]
                    if missing:
                        for name in missing:
                            ws.cell(row=1, column=ws.max_column + 1, value=name)
                            # Backfill blanks for existing data rows
                            for r in range(2, ws.max_row + 1):
                                ws.cell(row=r, column=ws.max_column, value="")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Tickets'
            ws.append(columns)

        # Prepare row
        row_dict = dict(ticket)
        if isinstance(row_dict.get('ai_response'), list):
            row_dict['ai_response'] = json.dumps(row_dict['ai_response'])
        if isinstance(row_dict.get('tags'), list):
            row_dict['tags'] = json.dumps(row_dict['tags'])
        row = [row_dict.get(col, '') for col in columns]

        ws.append(row)

        # Safe write with retries and atomic replace
        temp_path = excel_path + ".tmp"
        last_err = None
        for attempt in range(3):
            try:
                wb.save(temp_path)
                # Atomic replace minimizes lock windows
                os.replace(temp_path, excel_path)
                last_err = None
                break
            except PermissionError as pe:
                last_err = pe
                time.sleep(0.8)
            except Exception as e:
                last_err = e
                break
        if last_err:
            # Fallback: save to autosave file to avoid data loss when main is locked
            autosave_path = os.path.splitext(excel_path)[0] + "_autosave.xlsx"
            try:
                wb.save(autosave_path)
                st.warning(f"Main Excel is locked. Saved to autosave: {autosave_path}. Close Excel/OneDrive lock to resume writing to {excel_path}.")
            except Exception:
                raise last_err
        return True
    except Exception as e:
        st.error(f"Failed to write ticket to Excel: {str(e)}")
        return False

def save_all_tickets_to_excel(tickets: List[Dict], excel_path: str) -> bool:
    """Overwrite the Excel file with all tickets (used after edits)."""
    try:
        columns = [
            'ticket_id', 'customer_email', 'customer_name', 'issue_summary', 'detailed_issue',
            'category', 'priority', 'tags', 'status', 'created_date', 'created_time',
            'ai_response', 'confidence', 'solved', 'platform', 'contact_type', 'sentiment'
        ]
        rows = []
        for t in tickets:
            row = dict(t)
            if isinstance(row.get('ai_response'), list):
                row['ai_response'] = json.dumps(row['ai_response'])
            if isinstance(row.get('tags'), list):
                row['tags'] = json.dumps(row['tags'])
            rows.append([row.get(c, '') for c in columns])

        # Build a new workbook; if an old file exists, migrate headers to include any new columns
        wb = Workbook()
        ws = wb.active
        ws.title = 'Tickets'
        ws.append(columns)
        for r in rows:
            ws.append(r)

        temp_path = excel_path + ".tmp"
        last_err = None
        for attempt in range(3):
            try:
                wb.save(temp_path)
                os.replace(temp_path, excel_path)
                last_err = None
                break
            except PermissionError as pe:
                last_err = pe
                time.sleep(0.8)
            except Exception as e:
                last_err = e
                break
        if last_err:
            autosave_path = os.path.splitext(excel_path)[0] + "_autosave.xlsx"
            try:
                wb.save(autosave_path)
                st.warning(f"Main Excel is locked. Saved edited tickets to autosave: {autosave_path}.")
            except Exception:
                raise last_err
        return True
    except Exception as e:
        st.error(f"Failed to write all tickets to Excel: {str(e)}")
        return False

def sync_autosave_to_main(excel_path: str) -> bool:
    """If an autosave exists, merge it into the main Excel and remove autosave on success."""
    try:
        base, ext = os.path.splitext(excel_path)
        autosave_path = base + "_autosave" + ext
        if not os.path.exists(autosave_path):
            st.info("No autosave file found to sync.")
            return False

        # Load both files (if main missing, treat as empty)
        df_auto = pd.read_excel(autosave_path, engine='openpyxl')
        if os.path.exists(excel_path):
            df_main = pd.read_excel(excel_path, engine='openpyxl')
            # Merge on ticket_id without duplicates (autosave rows win)
            if 'ticket_id' in df_main.columns and 'ticket_id' in df_auto.columns:
                df_combined = pd.concat([df_main[~df_main['ticket_id'].isin(df_auto['ticket_id'])], df_auto], ignore_index=True)
            else:
                df_combined = pd.concat([df_main, df_auto], ignore_index=True)
        else:
            df_combined = df_auto

        # Write combined back using safe writer
        tickets = df_combined.to_dict(orient='records')
        ok = save_all_tickets_to_excel(tickets, excel_path)
        if ok:
            try:
                os.remove(autosave_path)
            except OSError:
                pass
            st.success("Autosave merged into main Excel.")
            return True
        return False
    except Exception as e:
        st.error(f"Failed to sync autosave: {str(e)}")
        return False

def create_sidebar():
    """Create the sidebar with settings."""
    st.sidebar.title("⚙️ Settings")
    
    # Excel autosave settings
    st.sidebar.subheader("📗 Excel Settings")
    st.session_state.excel_autosave = st.sidebar.checkbox("Autosave tickets to Excel", value=st.session_state.excel_autosave)
    st.session_state.excel_path = st.sidebar.text_input("Excel file path", value=st.session_state.excel_path)
    # Offer retry merge if autosave exists
    base, ext = os.path.splitext(st.session_state.excel_path)
    autosave_candidate = base + "_autosave" + ext
    if os.path.exists(autosave_candidate):
        if st.sidebar.button("🔁 Retry write (merge autosave)"):
            with st.spinner("Merging autosave into main Excel..."):
                sync_autosave_to_main(st.session_state.excel_path)
    if st.sidebar.button("📥 Load tickets from Excel"):
        with st.spinner("Loading tickets from Excel..."):
            xlsx_tickets = load_excel_data(st.session_state.excel_path)
            if xlsx_tickets:
                st.session_state.tickets = xlsx_tickets
                st.success(f"✅ Loaded {len(xlsx_tickets)} tickets from Excel")
            else:
                st.error("❌ No tickets loaded from Excel")

    # (Buttons removed per request)
    
    # Google Sheets Configuration
    st.sidebar.subheader("📊 Google Sheets Integration")
    # JSON upload removed per request; using existing/local auth methods only
    google_credentials = None
    sheet_id = st.sidebar.text_input(
        "Google Sheet ID",
        value="1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms",
        help="Enter your Google Sheet ID"
    )
    # Keep existing sheets client if already set; otherwise do nothing here
    
    # AI Configuration
    st.sidebar.subheader("🤖 AI Configuration")
    model_type = st.sidebar.selectbox(
        "AI Model",
        ["GPT-3.5", "GPT-4", "Gemini Pro", "Local Model"],
        index=0
    )
    
    confidence_threshold = st.sidebar.slider(
        "Confidence Threshold",
        min_value=0.0,
        max_value=1.0,
        value=0.8,
        step=0.05
    )
    
    # Google Search Configuration
    st.sidebar.subheader("🔍 Google Search Integration")
    st.sidebar.info("Enable Google search for enhanced chatbot responses")
    
    google_api_key = st.sidebar.text_input(
        "Google API Key",
        value=st.session_state.google_api_key or "",
        type="password",
        help="Enter your Google Custom Search API key"
    )
    
    google_cse_id = st.sidebar.text_input(
        "Google Custom Search Engine ID",
        value=st.session_state.google_cse_id or "",
        help="Enter your Google Custom Search Engine ID"
    )
    
    if google_api_key and google_cse_id:
        st.session_state.google_api_key = google_api_key
        st.session_state.google_cse_id = google_cse_id
        st.sidebar.success("✅ Google Search enabled!")
    else:
        st.sidebar.warning("⚠️ Google Search disabled - API key and CSE ID required")
    # Global search behavior controls
    st.session_state.search_always = st.sidebar.checkbox(
        "Always include Google results",
        value=bool(getattr(st.session_state, 'search_always', False))
    )
    st.session_state.search_conf_threshold = st.sidebar.slider(
        "Search trigger threshold (confidence)",
        min_value=0.0, max_value=1.0,
        value=float(getattr(st.session_state, 'search_conf_threshold', 0.7)), step=0.05,
        help="When knowledge confidence is below this, Google will be used to augment the answer"
    )
    
    if st.sidebar.button("🔍 Test Google Search"):
        if google_api_key and google_cse_id:
            try:
                import requests
                url = "https://www.googleapis.com/customsearch/v1"
                params = {
                    'key': google_api_key,
                    'cx': google_cse_id,
                    'q': 'test search',
                    'num': 1
                }
                response = requests.get(url, params=params, timeout=5)
                if response.status_code == 200:
                    st.sidebar.success("✅ Google Search API working!")
                else:
                    st.sidebar.error("❌ Google Search API failed")
            except Exception as e:
                st.sidebar.error(f"❌ Error: {str(e)}")
        else:
            st.sidebar.warning("Please enter API key and CSE ID first")
    # Email Notifications
    st.sidebar.subheader("✉️ Email Notifications")
    email_enabled = st.sidebar.checkbox("Enable email notifications", value=bool(getattr(st.session_state.notifier, 'enabled', False)))
    delivery_method = st.sidebar.selectbox(
        "Delivery method",
        ["smtp", "outlook", "eml", "mailto"],
        index=["smtp","outlook","eml","mailto"].index(getattr(getattr(st.session_state.notifier, 'email', None), 'delivery_method', 'smtp')) if getattr(getattr(st.session_state.notifier, 'email', None), 'delivery_method', 'smtp') in ["smtp","outlook","eml","mailto"] else 0
    )
    with st.sidebar.expander("Email Settings", expanded=False):
        smtp_host = st.text_input("SMTP host", value=getattr(getattr(st.session_state.notifier, 'email', None), 'smtp_host', '') or "")
        smtp_port = st.number_input("SMTP port", value=int(getattr(getattr(st.session_state.notifier, 'email', None), 'smtp_port', 587) or 587))
        smtp_user = st.text_input("SMTP username", value=getattr(getattr(st.session_state.notifier, 'email', None), 'username', '') or "")
        smtp_pass = st.text_input("SMTP password", type="password", value=getattr(getattr(st.session_state.notifier, 'email', None), 'password', '') or "")
        use_ssl = st.checkbox("Use SSL (implicit TLS)", value=bool(getattr(getattr(st.session_state.notifier, 'email', None), 'use_ssl', False)))
        use_tls = st.checkbox("Use STARTTLS", value=bool(getattr(getattr(st.session_state.notifier, 'email', None), 'use_tls', True)))
        verify = st.checkbox("Verify server certificate", value=bool(getattr(getattr(st.session_state.notifier, 'email', None), 'verify', True)))
        ca_file_uploader = st.file_uploader("Custom CA certificate (PEM)", type=['pem', 'crt', 'cer'])
        eml_dir = st.text_input("EML output directory", value=getattr(getattr(st.session_state.notifier, 'email', None), 'eml_out_dir', '') or "")
        sender = st.text_input("From email", value=getattr(getattr(st.session_state.notifier, 'email', None), 'sender', '') or "")
        recipient = st.text_input("To email", value=getattr(getattr(st.session_state.notifier, 'email', None), 'recipient', '') or "")

    email_cfg = None
    if smtp_host and sender and recipient:
        try:
            email_cfg = EmailConfig(
                smtp_host=smtp_host,
                smtp_port=int(smtp_port),
                username=smtp_user or None,
                password=smtp_pass or None,
                use_tls=bool(use_tls),
                use_ssl=bool(use_ssl),
                verify=bool(verify),
                ca_file=None,
                delivery_method=delivery_method,
                eml_out_dir=eml_dir or None,
                sender=sender,
                recipient=recipient,
            )
        except Exception:
            email_cfg = None
    # Persist CA file to temp if provided
    if email_cfg and ca_file_uploader is not None:
        try:
            ca_temp = 'smtp_ca.pem'
            with open(ca_temp, 'wb') as f:
                f.write(ca_file_uploader.getvalue())
            email_cfg.ca_file = ca_temp
        except Exception:
            pass
    st.session_state.notifier.update(enabled=email_enabled, email=email_cfg)

    if st.sidebar.button("✉️ Send Test Email"):
        ok = False
        try:
            ok = bool(st.session_state.notifier and st.session_state.notifier.send_test_email())
        except Exception as _e:
            ok = False
            st.session_state.notifier.last_error = str(_e)
        if ok:
            st.sidebar.success("Test email sent.")
        else:
            info = getattr(st.session_state.notifier, 'last_info', None) or ""
            err = getattr(st.session_state.notifier, 'last_error', None) or ""
            if info:
                st.sidebar.info(info)
            if err:
                st.sidebar.error(f"Failed to send test email: {err}")
            else:
                st.sidebar.error("Failed to send test email. Check SMTP settings.")

    
    # Knowledge Base & RAG
    st.sidebar.subheader("📚 Knowledge Base & RAG")
    
    # RAG Engine Status
    if hasattr(st.session_state.resolver, 'rag_engine') and st.session_state.resolver.rag_engine:
        stats = st.session_state.resolver.rag_engine.get_stats()
        st.sidebar.success(f"✅ RAG Active: {stats['total_documents']} docs")
        st.sidebar.caption(f"Model: {stats['model_name']}")
    else:
        st.sidebar.warning("⚠️ RAG Disabled - Using Keywords")
    
    if st.sidebar.button("🔄 Refresh Knowledge Base"):
        st.session_state.resolver.load_knowledge_base()
        # Rebuild RAG if available
        if hasattr(st.session_state.resolver, 'rag_engine') and st.session_state.resolver.rag_engine:
            try:
                documents = create_documents_from_knowledge_base(st.session_state.resolver.knowledge_base)
                st.session_state.resolver.rag_engine.rebuild_index(documents)
                st.sidebar.success("Knowledge base and RAG refreshed!")
            except Exception as e:
                st.sidebar.error(f"RAG refresh failed: {str(e)}")
        else:
            st.sidebar.success("Knowledge base refreshed!")
    
    # Toggle RAG
    if st.sidebar.button("🤖 Toggle RAG Engine"):
        if hasattr(st.session_state.resolver, 'use_rag'):
            st.session_state.resolver.use_rag = not st.session_state.resolver.use_rag
            if st.session_state.resolver.use_rag:
                st.sidebar.success("RAG enabled!")
            else:
                st.sidebar.info("RAG disabled - using keywords")
        else:
            st.sidebar.error("RAG not available")
    # Download KB as Excel
    try:
        kb = getattr(st.session_state.resolver, 'knowledge_base', {})
        if isinstance(kb, dict) and kb:
            # Prepare DataFrame
            kb_rows = []
            for key, data in kb.items():
                kb_rows.append({
                    'key': key,
                    'problem': data.get('problem', ''),
                    'keywords': ', '.join(data.get('keywords', [])),
                    'solutions': '\n'.join(data.get('solutions', [])),
                    'category': data.get('category', '')
                })
            if kb_rows:
                import pandas as _pd
                df_kb = _pd.DataFrame(kb_rows, columns=['key','problem','keywords','solutions','category'])
                bio = BytesIO()
                with _pd.ExcelWriter(bio, engine='openpyxl') as writer:
                    df_kb.to_excel(writer, index=False, sheet_name='KnowledgeBase')
                bio.seek(0)
                st.sidebar.download_button(
                    label="⬇ Download Knowledge Base (Excel)",
                    data=bio,
                    file_name="knowledge_base.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as _e:
        st.sidebar.error(f"KB export failed: {str(_e)}")
    
    return {
        'google_credentials': google_credentials,
        'sheet_id': sheet_id,
        'model_type': model_type,
        'confidence_threshold': confidence_threshold
    }

def display_main_header():
    """Display the main header."""
    st.markdown('<h1 class="main-header">🤖 AI Powered Knowledge Engine</h1>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Smart Support and Ticket Resolution</div>', unsafe_allow_html=True)
    st.markdown("---")

def create_query_resolution_tab():
    """Create the main query resolution tab."""
    st.header("💬 Customer Query Resolution")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("Customer Information")
        customer_email = st.text_input("Customer Email", placeholder="customer@example.com")
        customer_name = st.text_input("Customer Name", placeholder="John Doe")
        
        st.subheader("Issue Details")
        issue_summary = st.text_input("Issue Summary", placeholder="Brief description of the problem")
        detailed_issue = st.text_area("Detailed Issue Description", placeholder="Please provide more details about your issue...", height=100)
        
        if st.button("🔍 Get AI Response", type="primary"):
            if customer_email and issue_summary and detailed_issue:
                with st.spinner("Analyzing your query..."):
                    # Get AI response using resolver
                    query_response = st.session_state.resolver.resolve_query(detailed_issue)
                    # Temporarily store pending query context
                    pending_ticket_id = f"TK-{uuid.uuid4().hex[:12].upper()}"
                    st.session_state.pending_query = {
                        'customer_email': customer_email,
                        'customer_name': customer_name,
                        'issue_summary': issue_summary,
                        'detailed_issue': detailed_issue,
                        'query_response': query_response,
                        'pending_ticket_id': pending_ticket_id
                    }
                
            else:
                st.error("Please fill in all required fields")

    # If we have a pending AI response, display it with satisfaction buttons
    if st.session_state.pending_query is not None:
        pq = st.session_state.pending_query
        st.subheader("🤖 AI Response")
        # Show pending ticket id first
        if pq.get('pending_ticket_id'):
            st.success(f"Ticket ID: {pq['pending_ticket_id']}")
        
        # Show search method and confidence
        search_method = pq['query_response'].get('search_method', 'Unknown')
        confidence = pq['query_response']['confidence']
        category = pq['query_response']['category']
        
        if search_method == "RAG":
            st.info(f"🧠 RAG Search | Category: {category} | Confidence: {confidence:.2f}")
        elif search_method == "Keywords":
            st.info(f"🔍 Keyword Search | Category: {category} | Confidence: {confidence:.2f}")
        else:
            st.info(f"Category: {category} | Confidence: {confidence:.2f}")
        
        st.write("**Recommended Solutions:**")
        for i, solution in enumerate(pq['query_response']["solutions"], 1):
            st.write(f"{i}. {solution}")

        col_ok, col_not = st.columns(2)
        with col_ok:
            if st.button("✅ Satisfied - Close without ticket"):
                # Create a closed ticket (documentation of resolved query)
                response = dict(pq['query_response'])
                response['solved'] = True
                ticket = create_ticket(
                    pq['customer_email'], pq['customer_name'], pq['issue_summary'], pq['detailed_issue'],
                    response, status="Closed", solved=True, ticket_id_override=pq.get('pending_ticket_id')
                )
                st.session_state.last_created_ticket_id = ticket['ticket_id']
                st.session_state.last_created_ticket_ai = ticket['ai_response']
                st.session_state.flash_ticket_id = ticket['ticket_id']
                st.session_state.flash_ticket_status = ticket['status']
                st.success("Great! Marked as resolved and saved.")
                st.session_state.pending_query = None
                st.rerun()

        with col_not:
            if st.button("❌ Not satisfied - Create support ticket"):
                response = dict(pq['query_response'])
                response['solved'] = False
                ticket = create_ticket(
                    pq['customer_email'], pq['customer_name'], pq['issue_summary'], pq['detailed_issue'],
                    response, status="Open", solved=False, ticket_id_override=pq.get('pending_ticket_id')
                )
                st.session_state.last_created_ticket_id = ticket['ticket_id']
                st.session_state.last_created_ticket_ai = ticket['ai_response']
                st.session_state.flash_ticket_id = ticket['ticket_id']
                st.session_state.flash_ticket_status = ticket['status']
                st.warning(f"Ticket Created: {ticket['ticket_id']} - Our team will follow up.")
                st.session_state.pending_query = None
                st.rerun()
    
    with col2:
        st.subheader("📊 Quick Stats")
        total_tickets = len(st.session_state.tickets)
        solved_tickets = sum(1 for t in st.session_state.tickets if t["solved"])
        
        st.metric("Total Tickets", total_tickets)
        st.metric("Solved Today", solved_tickets)
        st.metric("Resolution Rate", f"{(solved_tickets/total_tickets*100):.1f}%" if total_tickets > 0 else "0%")
        
        # Last created ticket banner
        if st.session_state.last_created_ticket_id:
            st.markdown(f"**Last Ticket:** {st.session_state.last_created_ticket_id}")
            if st.session_state.last_created_ticket_ai:
                st.write("**AI Responses:**")
                for i, resp in enumerate(st.session_state.last_created_ticket_ai, 1):
                    st.write(f"{i}. {resp}")
        # One-time flash banner right after creation
        if st.session_state.flash_ticket_id:
            st.success(f"🎫 Ticket Created: {st.session_state.flash_ticket_id} | Status: {st.session_state.flash_ticket_status}")
            # Clear flash so it doesn't persist across reruns
            st.session_state.flash_ticket_id = None
            st.session_state.flash_ticket_status = None

        # Recent tickets (show only the latest)
        if st.session_state.tickets:
            st.subheader("📋 Recent Tickets")
            for ticket in st.session_state.tickets[-1:]:
                with st.container():
                    st.write(f"**{ticket['ticket_id']}**")
                    st.write(f"*{ticket['issue_summary']}*")
                    st.write(f"Status: {ticket['status']}")

def create_ticket(customer_email: str, customer_name: str, issue_summary: str, detailed_issue: str, query_response: Dict, status: str = "Open", solved: bool | None = None, ticket_id_override: str | None = None) -> Dict:
    """Create a new support ticket using categorizer and tagger."""
    # Use high-entropy unique ID to avoid collisions in the same second
    ticket_id = ticket_id_override or f"TK-{uuid.uuid4().hex[:12].upper()}"
    
    # Use categorizer to categorize the ticket
    category_result = st.session_state.categorizer.categorize(detailed_issue)
    
    # Use tagger to extract tags
    tags = st.session_state.tagger.extract_tags(detailed_issue)
    
    # Decide solved flag
    solved_flag = query_response["solved"] if solved is None else solved

    ticket = {
        "ticket_id": ticket_id,
        "customer_email": customer_email,
        "customer_name": customer_name,
        "issue_summary": issue_summary,
        "detailed_issue": detailed_issue,
        "category": category_result["category"],
        "priority": category_result["priority"],
        "tags": tags,
        "status": status,
        "created_date": datetime.now().strftime("%Y-%m-%d"),
        "created_time": datetime.now().strftime("%H:%M:%S"),
        "ai_response": query_response["solutions"],
        "confidence": query_response["confidence"],
        "solved": solved_flag,
        "platform": "Web",
        "contact_type": "Web Form",
        "sentiment": compute_sentiment_label(detailed_issue or issue_summary)
    }
    
    st.session_state.tickets.append(ticket)
    # Persist to Excel automatically
    if st.session_state.excel_autosave:
        success_excel = save_ticket_to_excel(ticket, st.session_state.excel_path)
        if success_excel:
            pass
    # Notify external systems
    try:
        if 'notifier' in st.session_state and st.session_state.notifier:
            st.session_state.notifier.send_ticket_created(ticket)
            # Category volume alert: if category count > 3, send alert
            try:
                cat = ticket.get('category', 'General')
                count_cat = sum(1 for t in st.session_state.tickets if t.get('category') == cat)
                if count_cat >= 4:
                    # Send one alert per threshold crossing; simple debounce via last_info text
                    st.session_state.notifier.send_category_threshold_alert(cat, count_cat)
            except Exception:
                pass
    except Exception:
        pass
    return ticket

def create_ticket_management_tab():
    """Create ticket management tab."""
    st.header("🎫 Ticket Management")
    
    if st.session_state.tickets:
        # Filters
        col1, col2, col3 = st.columns(3)
        with col1:
            status_filter = st.selectbox("Filter by Status", ["All", "Open", "Closed", "In Progress"])
        with col2:
            priority_filter = st.selectbox("Filter by Priority", ["All", "High", "Medium", "Low"])
        with col3:
            category_filter = st.selectbox("Filter by Category", ["All"] + list(set(t["category"] for t in st.session_state.tickets)))
        
        # Cleanup tools
        st.markdown("---")
        st.subheader("🧹 Cleanup Tools")
        c1, c2, c3 = st.columns([1,1,2])
        with c1:
            confirm_cleanup = st.checkbox("Confirm cleanup", key="confirm_cleanup")
        with c2:
            if st.button("Delete CLOSED tickets"):
                if confirm_cleanup:
                    before = len(st.session_state.tickets)
                    st.session_state.tickets = [t for t in st.session_state.tickets if str(t.get('status','')).strip().lower() != 'closed']
                    after = len(st.session_state.tickets)
                    if st.session_state.excel_autosave:
                        save_all_tickets_to_excel(st.session_state.tickets, st.session_state.excel_path)
                    st.success(f"Removed {before - after} closed tickets")
                    st.rerun()
                else:
                    st.warning("Please check 'Confirm cleanup' before deleting.")
        with c3:
            if st.button("Delete tickets with empty summary/details"):
                if confirm_cleanup:
                    before = len(st.session_state.tickets)
                    def _is_meaningful(s: str) -> bool:
                        try:
                            text = str(s or '').strip().lower()
                            if not text or len(text) < 3:
                                return False
                            placeholders = {"test","dummy","sample","na","n/a","none","-","--","?","asd","asdf"}
                            if text in placeholders:
                                return False
                            import re as _re
                            if not _re.search(r"[a-z0-9]", text):
                                return False
                            return True
                        except Exception:
                            return False
                    st.session_state.tickets = [
                        t for t in st.session_state.tickets
                        if _is_meaningful(t.get('issue_summary')) or _is_meaningful(t.get('detailed_issue'))
                    ]
                    after = len(st.session_state.tickets)
                    if st.session_state.excel_autosave:
                        save_all_tickets_to_excel(st.session_state.tickets, st.session_state.excel_path)
                    st.success(f"Removed {before - after} empty tickets")
                    st.rerun()
                else:
                    st.warning("Please check 'Confirm cleanup' before deleting.")
        
        # Filter tickets
        filtered_tickets = st.session_state.tickets
        if status_filter != "All":
            filtered_tickets = [t for t in filtered_tickets if t["status"] == status_filter]
        if priority_filter != "All":
            filtered_tickets = [t for t in filtered_tickets if t["priority"] == priority_filter]
        if category_filter != "All":
            filtered_tickets = [t for t in filtered_tickets if t["category"] == category_filter]
        
        # Display tickets
        for idx, ticket in enumerate(filtered_tickets):
            with st.expander(f"🎫 {ticket['ticket_id']} - {ticket['issue_summary']}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write(f"**Customer:** {ticket['customer_name']} ({ticket['customer_email']})")
                    st.write(f"**Category:** {ticket['category']}")
                    st.write(f"**Priority:** {ticket['priority']}")
                    st.write(f"**Status:** {ticket['status']}")
                    st.write(f"**Created:** {ticket['created_date']} {ticket['created_time']}")
                
                with col2:
                    st.write(f"**Confidence:** {ticket['confidence']:.2f}")
                    st.write(f"**Solved:** {'✅ Yes' if ticket['solved'] else '❌ No'}")
                    st.write(f"**Platform:** {ticket['platform']}")
                    st.write(f"**Tags:** {', '.join(ticket['tags'])}")
                    st.write(f"**Sentiment:** {ticket.get('sentiment','Neutral')}")
                
                st.write("**Issue Description:**")
                st.write(ticket['detailed_issue'])
                
                st.write("**AI Response:**")
                for response in ticket['ai_response']:
                    st.write(f"• {response}")
                
                # Action buttons
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    if st.button(f"✅ Mark Solved", key=f"solve_{ticket['ticket_id']}_{idx}"):
                        ticket['status'] = "Closed"
                        ticket['solved'] = True
                        if st.session_state.excel_autosave:
                            save_all_tickets_to_excel(st.session_state.tickets, st.session_state.excel_path)
                            # Attempt inline merge if autosave exists (e.g., when Excel was locked)
                            base, ext = os.path.splitext(st.session_state.excel_path)
                            autosave_candidate = base + "_autosave" + ext
                            if os.path.exists(autosave_candidate):
                                sync_autosave_to_main(st.session_state.excel_path)
                        st.success("Ticket marked as solved!")
                        st.rerun()
                
                with col2:
                    if st.button(f"🔄 In Progress", key=f"progress_{ticket['ticket_id']}_{idx}"):
                        ticket['status'] = "In Progress"
                        if st.session_state.excel_autosave:
                            save_all_tickets_to_excel(st.session_state.tickets, st.session_state.excel_path)
                            base, ext = os.path.splitext(st.session_state.excel_path)
                            autosave_candidate = base + "_autosave" + ext
                            if os.path.exists(autosave_candidate):
                                sync_autosave_to_main(st.session_state.excel_path)
                        st.success("Ticket status updated!")
                        st.rerun()
                
                with col3:
                    if st.button(f"↩️ Reopen (Unsolve)", key=f"reopen_{ticket['ticket_id']}_{idx}"):
                        ticket['status'] = "Open"
                        ticket['solved'] = False
                        if st.session_state.excel_autosave:
                            save_all_tickets_to_excel(st.session_state.tickets, st.session_state.excel_path)
                            base, ext = os.path.splitext(st.session_state.excel_path)
                            autosave_candidate = base + "_autosave" + ext
                            if os.path.exists(autosave_candidate):
                                sync_autosave_to_main(st.session_state.excel_path)
                        st.success("Ticket reopened and marked as unsolved!")
                        st.rerun()

                with col4:
                    edit_key = f"edit_mode_{ticket['ticket_id']}"
                    if edit_key not in st.session_state:
                        st.session_state[edit_key] = False
                    if st.button(f"📝 Edit", key=f"edit_{ticket['ticket_id']}_{idx}"):
                        st.session_state[edit_key] = not st.session_state[edit_key]

                # Row for destructive action
                del_col, _sp = st.columns([1,3])
                with del_col:
                    if st.button("🗑️ Delete", key=f"delete_{ticket['ticket_id']}_{idx}"):
                        to_delete_id = str(ticket.get('ticket_id'))
                        st.session_state.tickets = [t for t in st.session_state.tickets if str(t.get('ticket_id')) != to_delete_id]
                        if st.session_state.excel_autosave:
                            save_all_tickets_to_excel(st.session_state.tickets, st.session_state.excel_path)
                        st.warning(f"Deleted ticket {ticket['ticket_id']}")
                        st.rerun()

                if st.session_state.get(edit_key, False):
                    st.markdown("---")
                    st.subheader("Edit Ticket")
                    e_col1, e_col2 = st.columns(2)
                    with e_col1:
                        new_summary = st.text_input("Issue Summary", value=ticket['issue_summary'], key=f"sum_{ticket['ticket_id']}_{idx}")
                        new_status = st.selectbox("Status", ["Open", "In Progress", "Closed"], index=["Open","In Progress","Closed"].index(ticket['status']), key=f"stat_{ticket['ticket_id']}_{idx}")
                        new_priority = st.selectbox("Priority", ["Low","Medium","High","Critical"], index=["Low","Medium","High","Critical"].index(ticket['priority']) if ticket['priority'] in ["Low","Medium","High","Critical"] else 1, key=f"prio_{ticket['ticket_id']}_{idx}")
                        new_solved = st.checkbox("Solved", value=ticket['solved'], key=f"solv_{ticket['ticket_id']}_{idx}")
                    with e_col2:
                        new_category = st.text_input("Category", value=ticket['category'], key=f"cat_{ticket['ticket_id']}_{idx}")
                        new_tags_str = st.text_input("Tags (comma separated)", value=", ".join(ticket['tags']), key=f"tags_{ticket['ticket_id']}_{idx}")
                    new_detail = st.text_area("Detailed Issue Description", value=ticket['detailed_issue'], key=f"det_{ticket['ticket_id']}_{idx}")

                    b1, b2 = st.columns(2)
                    with b1:
                        if st.button("💾 Save Changes", key=f"save_{ticket['ticket_id']}_{idx}"):
                            ticket['issue_summary'] = new_summary
                            ticket['detailed_issue'] = new_detail
                            ticket['status'] = new_status
                            ticket['priority'] = new_priority
                            ticket['category'] = new_category
                            ticket['solved'] = bool(new_solved)
                            ticket['tags'] = [t.strip() for t in new_tags_str.split(',') if t.strip()]
                            if st.session_state.excel_autosave:
                                ok = save_all_tickets_to_excel(st.session_state.tickets, st.session_state.excel_path)
                                if ok:
                                    pass
                                base, ext = os.path.splitext(st.session_state.excel_path)
                                autosave_candidate = base + "_autosave" + ext
                                if os.path.exists(autosave_candidate):
                                    sync_autosave_to_main(st.session_state.excel_path)
                            st.session_state[edit_key] = False
                            # Notify on ticket update
                            try:
                                if 'notifier' in st.session_state and st.session_state.notifier:
                                    st.session_state.notifier.send_ticket_updated(ticket)
                            except Exception:
                                pass
                            st.success("Ticket updated.")
                            st.rerun()
                    with b2:
                        if st.button("✖ Cancel", key=f"cancel_{ticket['ticket_id']}_{idx}"):
                            st.session_state[edit_key] = False
    else:
        st.info("No tickets created yet. Go to 'Query Resolution' tab to create tickets.")

def create_google_sheets_tab():
    """Create Google Sheets integration tab."""
    st.header("📊 Data Integration & Export")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📥 Import Data")
        
        # CSV Import
        st.write("**From CSV File:**")
        if st.button("📊 Load Sample Data from CSV", type="primary"):
            with st.spinner("Loading sample data from CSV..."):
                csv_tickets = load_csv_data()
                if csv_tickets:
                    st.session_state.tickets = csv_tickets
                    st.session_state.csv_loaded = True
                    st.success(f"✅ Loaded {len(csv_tickets)} tickets from CSV!")
                else:
                    st.error("❌ Failed to load CSV data")
        
        # Google Sheets Import
        st.write("**From Google Sheets:**")
        if st.button("📊 Load Data from Google Sheets"):
            if st.session_state.sheets_client.sheet_id:
                with st.spinner("Loading data from Google Sheets..."):
                    tickets = st.session_state.sheets_client.read_tickets_from_sheet()
                    
                    if tickets:
                        st.session_state.google_sheet_data = tickets
                        st.success(f"✅ Loaded {len(tickets)} records from Google Sheets")
                        
                        # Display data
                        df = pd.DataFrame(tickets)
                        st.dataframe(df)
                    else:
                        st.error("❌ No data found in Google Sheets")
            else:
                st.error("❌ Please configure Google Sheet ID in settings")
    
    with col2:
        st.subheader("📤 Export Data")
        
        if st.session_state.tickets:
            # CSV Export
            st.write("**Export to CSV:**")
            if st.button("💾 Export Tickets to CSV", type="primary"):
                try:
                    df = pd.DataFrame(st.session_state.tickets)
                    csv_data = df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download CSV",
                        data=csv_data,
                        file_name=f"tickets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                    st.success("✅ CSV export ready for download!")
                except Exception as e:
                    st.error(f"❌ Error creating CSV export: {str(e)}")
            
            # Google Sheets Export
            st.write("**Export to Google Sheets:**")
            if st.button("💾 Export Tickets to Google Sheets"):
                with st.spinner("Exporting tickets to Google Sheets..."):
                    success = st.session_state.sheets_client.write_tickets_to_sheet(st.session_state.tickets)
                    
                    if success:
                        st.success("✅ Tickets exported to Google Sheets successfully!")
                    else:
                        st.error("❌ Failed to export tickets to Google Sheets")
        else:
            st.info("No tickets to export. Create some tickets first!")

def create_analytics_tab():
    """Create analytics and reporting tab."""
    st.header("📈 Analytics & Reports")
    
    if st.session_state.tickets:
        # Metrics
        total_tickets = len(st.session_state.tickets)
        solved_tickets = sum(1 for t in st.session_state.tickets if t["solved"])
        open_tickets = sum(1 for t in st.session_state.tickets if t["status"] == "Open")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Tickets", total_tickets)
        with col2:
            st.metric("Solved Tickets", solved_tickets)
        with col3:
            st.metric("Open Tickets", open_tickets)
        with col4:
            st.metric("Resolution Rate", f"{(solved_tickets/total_tickets*100):.1f}%")
        
        # Charts
        col1, col2 = st.columns(2)
        
        with col1:
            # Category distribution
            category_counts = {}
            for ticket in st.session_state.tickets:
                category = ticket["category"]
                category_counts[category] = category_counts.get(category, 0) + 1
            
            if category_counts:
                fig_cat = px.pie(
                    values=list(category_counts.values()),
                    names=list(category_counts.keys()),
                    title="Tickets by Category"
                )
                st.plotly_chart(fig_cat, use_container_width=True)
        
        with col2:
            # Priority distribution
            priority_counts = {}
            for ticket in st.session_state.tickets:
                priority = ticket["priority"]
                priority_counts[priority] = priority_counts.get(priority, 0) + 1
            
            if priority_counts:
                fig_priority = px.bar(
                    x=list(priority_counts.keys()),
                    y=list(priority_counts.values()),
                    title="Tickets by Priority",
                    color=list(priority_counts.keys()),
                    color_discrete_sequence=px.colors.qualitative.Set3
                )
                st.plotly_chart(fig_priority, use_container_width=True)
        
        # Daily ticket trends
        if len(st.session_state.tickets) > 1:
            st.subheader("📊 Daily Ticket Trends")
            daily_counts = {}
            for ticket in st.session_state.tickets:
                date = ticket["created_date"]
                daily_counts[date] = daily_counts.get(date, 0) + 1
            
            if daily_counts:
                df_daily = pd.DataFrame(list(daily_counts.items()), columns=['Date', 'Tickets'])
                fig_trend = px.line(df_daily, x='Date', y='Tickets', title='Tickets Created Over Time')
                st.plotly_chart(fig_trend, use_container_width=True)

        # ---------- Visual Enhancements & Extra Insights (additive) ----------
        st.markdown("---")
        st.subheader("🎨 Visual Insights")

        # Sentiment distribution (donut)
        sentiments = [t.get('sentiment', 'Neutral') for t in st.session_state.tickets]
        if sentiments:
            sent_counts = {}
            for s in sentiments:
                sent_counts[s or 'Neutral'] = sent_counts.get(s or 'Neutral', 0) + 1
            df_sent = pd.DataFrame({
                'Sentiment': list(sent_counts.keys()),
                'Count': list(sent_counts.values())
            })
            fig_sent = px.pie(df_sent, names='Sentiment', values='Count', hole=0.45,
                               title='Sentiment Distribution',
                               color='Sentiment',
                               color_discrete_map={'Positive':'#10B981','Neutral':'#6B7280','Negative':'#EF4444'})
            st.plotly_chart(fig_sent, use_container_width=True)

        # Status funnel (bar sorted)
        status_counts = {}
        for t in st.session_state.tickets:
            stx = t.get('status', 'Open')
            status_counts[stx] = status_counts.get(stx, 0) + 1
        if status_counts:
            df_status = pd.DataFrame(sorted(status_counts.items(), key=lambda x: x[1], reverse=True),
                                     columns=['Status','Count'])
            fig_status = px.bar(df_status, x='Status', y='Count', title='Ticket Status Overview',
                                color='Status', color_discrete_sequence=px.colors.qualitative.Set2)
            st.plotly_chart(fig_status, use_container_width=True)

        # Confidence distribution (histogram)
        confidences = [float(t.get('confidence', 0)) for t in st.session_state.tickets]
        if confidences:
            fig_conf = px.histogram(x=confidences, nbins=10, title='AI Confidence Distribution',
                                    labels={'x':'Confidence','y':'Count'}, color_discrete_sequence=['#6366F1'])
            st.plotly_chart(fig_conf, use_container_width=True)

        # Top tags (bar)
        tag_counts = {}
        for t in st.session_state.tickets:
            if isinstance(t.get('tags'), list):
                for tag in t['tags']:
                    tag_counts[tag] = tag_counts.get(tag, 0) + 1
        if tag_counts:
            df_tags = pd.DataFrame(sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:15],
                                   columns=['Tag','Count'])
            fig_tags = px.bar(df_tags, x='Tag', y='Count', title='Top Tags',
                              color='Count', color_continuous_scale='Tealgrn')
            st.plotly_chart(fig_tags, use_container_width=True)

        # Category vs Sentiment (stacked)
        try:
            rows = []
            for t in st.session_state.tickets:
                rows.append({
                    'Category': t.get('category', 'General'),
                    'Sentiment': t.get('sentiment', 'Neutral') or 'Neutral'
                })
            if rows:
                df_cs = pd.DataFrame(rows)
                df_cs['Count'] = 1
                fig_cs = px.bar(df_cs, x='Category', y='Count', color='Sentiment', barmode='stack',
                                title='Sentiment by Category', color_discrete_map={'Positive':'#10B981','Neutral':'#6B7280','Negative':'#EF4444'})
                st.plotly_chart(fig_cs, use_container_width=True)
        except Exception:
            pass

        # Priority-Status matrix (stacked)
        try:
            rows = []
            for t in st.session_state.tickets:
                rows.append({
                    'Priority': t.get('priority', 'Medium'),
                    'Status': t.get('status', 'Open')
                })
            if rows:
                df_ps = pd.DataFrame(rows)
                df_ps['Count'] = 1
                fig_ps = px.bar(df_ps, x='Priority', y='Count', color='Status', barmode='stack',
                                title='Workload by Priority and Status', color_discrete_sequence=px.colors.qualitative.Set3)
                st.plotly_chart(fig_ps, use_container_width=True)
        except Exception:
            pass

        # Category footprint treemap removed per request

        # Downloads: export analytics aggregates
        with st.expander('⬇ Download Analytics Data', expanded=False):
            try:
                # Sentiment export
                if sentiments:
                    st.download_button('Download Sentiment Counts (CSV)',
                        data=pd.DataFrame({'Sentiment': list(sent_counts.keys()), 'Count': list(sent_counts.values())}).to_csv(index=False),
                        file_name='sentiment_counts.csv', mime='text/csv')
                # Status export
                if status_counts:
                    st.download_button('Download Status Counts (CSV)',
                        data=df_status.to_csv(index=False), file_name='status_counts.csv', mime='text/csv')
                # Tags export
                if tag_counts:
                    st.download_button('Download Top Tags (CSV)',
                        data=df_tags.to_csv(index=False), file_name='top_tags.csv', mime='text/csv')
            except Exception:
                pass

        # Volume heatmap by Day-of-Week vs Hour (if we have times)
        try:
            dow_hour = []
            for t in st.session_state.tickets:
                d = t.get('created_date')
                h = t.get('created_time')
                if d and h:
                    try:
                        dt = pd.to_datetime(f"{d} {h}")
                    except Exception:
                        dt = pd.to_datetime(str(d))
                    dow_hour.append({'dow': dt.day_name(), 'hour': dt.hour})
            if dow_hour:
                df_dh = pd.DataFrame(dow_hour)
                pivot = df_dh.pivot_table(index='dow', columns='hour', aggfunc=len, fill_value=0)
                # Order days
                day_order = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
                pivot = pivot.reindex(day_order)
                fig_heat = px.imshow(pivot, aspect='auto', color_continuous_scale='Blues',
                                     title='Ticket Volume Heatmap (Day vs Hour)')
                st.plotly_chart(fig_heat, use_container_width=True)
        except Exception:
            pass

        # --- Knowledge coverage analytics ---
        st.subheader("📚 Knowledge Base Coverage")
        # Track which KB articles are referenced via resolver responses
        kb_usage: Dict[str, int] = {}
        # We will infer usage from the last created ticket AI responses when possible
        # and from matched kb_key recorded in pending_query/last_created_ticket_ai or stored tickets
        for t in st.session_state.tickets:
            kb_key = None
            # If we stored kb_key in ticket tags previously, attempt to read (optional)
            if isinstance(t.get('tags'), list):
                for tag in t['tags']:
                    if str(tag).startswith('kb:'):
                        kb_key = str(tag)[3:]
                        break
            if not kb_key:
                # fallback: derive a simple hash from first solution to group usage (best-effort)
                if isinstance(t.get('ai_response'), list) and t['ai_response']:
                    kb_key = f"auto:{hash(t['ai_response'][0]) % 1000000}"
            if kb_key:
                kb_usage[kb_key] = kb_usage.get(kb_key, 0) + 1

        if kb_usage:
            df_kb = pd.DataFrame(
                sorted(kb_usage.items(), key=lambda x: x[1], reverse=True),
                columns=['Article', 'References']
            )
            st.plotly_chart(px.bar(df_kb.head(10), x='Article', y='References', title='Top Referenced KB Articles'), use_container_width=True)

            # Identify low-coverage categories (few/no references)
            kb = getattr(st.session_state.resolver, 'knowledge_base', {})
            all_articles = set(kb.keys()) if isinstance(kb, dict) else set()
            referenced = set(a for a in kb_usage.keys() if isinstance(a, str) and not a.startswith('auto:'))
            unused = list(all_articles - referenced)
            st.info(f"Unused/low-reference KB articles: {len(unused)}")
            if unused:
                st.dataframe(pd.DataFrame(unused, columns=['Unused Articles']).head(20))

        # --- Automated alert for low coverage support areas ---
        st.subheader("⚠️ Low-Coverage Alerts")
        # Compute category coverage: proportion of tickets that produced AI responses
        category_totals: Dict[str, int] = {}
        category_with_ai: Dict[str, int] = {}
        for t in st.session_state.tickets:
            cat = t.get('category', 'General')
            category_totals[cat] = category_totals.get(cat, 0) + 1
            has_ai = 1 if (isinstance(t.get('ai_response'), list) and len(t['ai_response']) > 0) else 0
            category_with_ai[cat] = category_with_ai.get(cat, 0) + has_ai

        low_coverage_rows = []
        coverage_rows_all = []
        for cat, total in category_totals.items():
            covered = category_with_ai.get(cat, 0)
            coverage = covered / total if total else 0.0
            coverage_class = 'Low' if coverage < 0.4 else 'High'
            coverage_rows_all.append({'Category': cat, 'Tickets': total, 'Covered': covered, 'CoveragePct': round(coverage * 100, 2), 'CoverageClass': coverage_class})
            if coverage < 0.4 and total >= 3:  # alert threshold and minimum volume
                low_coverage_rows.append({
                    'Category': cat,
                    'Tickets': total,
                    'Coverage': f"{coverage:.0%}",
                    'Coverage Class': coverage_class,
                    'Suggestion': 'Expand KB content for this category'
                })

        if low_coverage_rows:
            st.error("Low coverage detected in some support areas")
            st.dataframe(pd.DataFrame(low_coverage_rows))
        else:
            st.success("No low-coverage alerts. KB coverage looks healthy.")

        # Coverage chart for all categories (visual)
        if coverage_rows_all:
            try:
                df_cov = pd.DataFrame(coverage_rows_all)
                df_cov = df_cov.sort_values('CoveragePct', ascending=True)
                fig_cov = px.bar(
                    df_cov,
                    x='CoveragePct', y='Category', orientation='h',
                    color='CoverageClass', color_discrete_map={'High':'#10B981','Low':'#EF4444'},
                    title='Knowledge Coverage by Category (%)',
                    labels={'CoveragePct':'Coverage %','Category':'Category','CoverageClass':'Class'}
                )
                # Add threshold line at 40%
                fig_cov.add_vline(x=40, line_dash='dash', line_color='red')
                st.plotly_chart(fig_cov, use_container_width=True)

                # Summary counts by class
                class_counts = df_cov['CoverageClass'].value_counts().to_dict()
                st.caption(f"Coverage classes — High: {class_counts.get('High',0)} | Low: {class_counts.get('Low',0)} (threshold 40%)")
            except Exception:
                pass
    else:
        st.info("No tickets available for analytics. Create some tickets first!")

def create_chatbot_tab():
    """Create the interactive chatbot tab."""
    st.header("🤖 AI Chatbot Support")
    
    # Initialize chatbot if not already done
    if st.session_state.chatbot is None:
        st.session_state.chatbot = Chatbot(
            resolver=st.session_state.resolver,
            categorizer=st.session_state.categorizer,
            tagger=st.session_state.tagger,
            google_api_key=st.session_state.google_api_key,
            google_cse_id=st.session_state.google_cse_id
        )
        st.session_state.chatbot_initialized = True
    
    # Restore conversation history if available
    if st.session_state.chat_history and not st.session_state.chatbot_initialized:
        try:
            st.session_state.chatbot.import_conversation(st.session_state.chat_history)
            st.session_state.chatbot_initialized = True
        except Exception as e:
            st.warning(f"Could not restore chat history: {str(e)}")
            st.session_state.chat_history = []
    
    # Update Google search credentials if they changed
    if st.session_state.chatbot and (
        st.session_state.chatbot.google_api_key != st.session_state.google_api_key or
        st.session_state.chatbot.google_cse_id != st.session_state.google_cse_id
    ):
        st.session_state.chatbot.google_api_key = st.session_state.google_api_key
        st.session_state.chatbot.google_cse_id = st.session_state.google_cse_id
        st.session_state.chatbot.search_enabled = bool(st.session_state.google_api_key and st.session_state.google_cse_id)
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("💬 Chat with AI Assistant")
        
        # Chat interface
        chat_container = st.container()
        
        # Display chat history
        with chat_container:
            if st.session_state.chatbot.conversation_history:
                for message in st.session_state.chatbot.conversation_history:
                    if message.role == "user":
                        with st.chat_message("user"):
                            st.write(message.content)
                    else:
                        with st.chat_message("assistant"):
                            st.write(message.content)
            else:
                st.info("👋 Hello! I'm your AI support assistant. How can I help you today?")
        
        # User input
        st.markdown("---")
        user_input = st.text_input(
            "Type your message here...",
            placeholder="Ask me anything about your issue or problem",
            key="chatbot_input"
        )
        
        col_send, col_clear = st.columns([1, 1])
        
        with col_send:
            if st.button("💬 Send Message", type="primary", key="send_chat"):
                if user_input.strip():
                    with st.spinner("AI is thinking..."):
                        # Process user message
                        response = st.session_state.chatbot.process_user_message(user_input.strip())
                        
                        # Store in session state for persistence
                        st.session_state.chat_history = st.session_state.chatbot.export_conversation()
                        
                        # Show success message
                        st.success("Message sent! Check the chat above.")
                        st.rerun()
                else:
                    st.warning("Please enter a message.")
        
        with col_clear:
            if st.button("🗑️ Clear Chat", key="clear_chat"):
                st.session_state.chatbot.clear_conversation()
                st.session_state.chat_history = []
                st.success("Chat cleared!")
                st.rerun()
    
    with col2:
        st.subheader("📊 Chat Statistics")
        
        # Chat summary
        summary = st.session_state.chatbot.get_conversation_summary()
        st.metric("Messages", summary.get("message_count", 0))
        st.metric("Duration", summary.get("conversation_duration", "0 minutes"))
        
        # Search status
        st.subheader("🔍 Search Status")
        if st.session_state.chatbot.search_enabled:
            st.success("✅ Google Search enabled")
        else:
            st.warning("⚠️ Google Search disabled")
            st.caption("Configure API key and CSE ID in sidebar to enable")
        
        # Current context
        if st.session_state.chatbot.current_context:
            st.subheader("🎯 Current Context")
            context = st.session_state.chatbot.current_context
            if context.get('category'):
                st.info(f"**Category:** {context['category']}")
            if context.get('confidence'):
                st.info(f"**Confidence:** {context['confidence']:.1%}")
        
        # Quick actions
        st.subheader("⚡ Quick Actions")
        
        if st.button("🎫 Create Ticket from Chat", key="create_ticket_from_chat"):
            if st.session_state.chatbot.conversation_history:
                # Get the last user message and AI response
                last_user_msg = None
                last_ai_msg = None
                
                for msg in reversed(st.session_state.chatbot.conversation_history):
                    if msg.role == "user" and not last_user_msg:
                        last_user_msg = msg.content
                    elif msg.role == "assistant" and not last_ai_msg:
                        last_ai_msg = msg.content
                    if last_user_msg and last_ai_msg:
                        break
                
                if last_user_msg:
                    # Create a ticket with the chat context
                    ticket_id = f"TK-{uuid.uuid4().hex[:12].upper()}"
                    
                    # Use the chatbot's context for ticket creation
                    context = st.session_state.chatbot.current_context
                    query_response = {
                        "solutions": [last_ai_msg] if last_ai_msg else ["Chat conversation recorded"],
                        "confidence": context.get('confidence', 0.5),
                        "category": context.get('category', 'Chat Support'),
                        "solved": False
                    }
                    
                    ticket = create_ticket(
                        customer_email="chat@support.com",
                        customer_name="Chat User",
                        issue_summary=last_user_msg[:100] + "..." if len(last_user_msg) > 100 else last_user_msg,
                        detailed_issue=last_user_msg,
                        query_response=query_response,
                        status="Open",
                        solved=False,
                        ticket_id_override=ticket_id
                    )
                    
                    st.success(f"🎫 Ticket created: {ticket_id}")
                    st.info("The chat conversation has been converted to a support ticket.")
                else:
                    st.warning("No conversation to convert to ticket.")
            else:
                st.warning("No chat history available.")
        
        if st.button("📋 Export Chat History", key="export_chat"):
            if st.session_state.chatbot.conversation_history:
                chat_data = st.session_state.chatbot.export_conversation()
                chat_json = json.dumps(chat_data, indent=2)
                
                st.download_button(
                    label="📥 Download Chat History",
                    data=chat_json,
                    file_name=f"chat_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json"
                )
            else:
                st.warning("No chat history to export.")
        
        # Suggested questions
        st.subheader("💡 Suggested Questions")
        suggested_questions = [
            "My phone screen is cracked, what should I do?",
            "I can't log into my account",
            "My phone is not charging properly",
            "I'm having network connectivity issues",
            "My phone is running slowly"
        ]
        
        for i, question in enumerate(suggested_questions):
            if st.button(f"💬 {question}", key=f"suggested_{i}"):
                # Process the suggested question
                with st.spinner("AI is thinking..."):
                    response = st.session_state.chatbot.process_user_message(question)
                    st.session_state.chat_history = st.session_state.chatbot.export_conversation()
                    st.success("Question sent! Check the chat above.")
                    st.rerun()
        
        # Helpful tips
        st.subheader("💡 Tips")
        st.info("""
        **How to use the chatbot:**
        - Ask specific questions about your issue
        - Provide details about what you're experiencing
        - Ask for clarification on solutions
        - Request additional help or alternatives
        """)

def main():
    """Main application function."""
    initialize_session_state()
    # --- Simple login gate ---
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    if 'auth_user' not in st.session_state:
        st.session_state.auth_user = None

    app_user = os.getenv('APP_USERNAME', '')
    app_pass = os.getenv('APP_PASSWORD', '')

    if not st.session_state.authenticated:
        # Background image for login screen
        try:
            img_path = os.path.join(os.getcwd(), "img.jpg")
            if os.path.exists(img_path):
                with open(img_path, "rb") as _bgf:
                    _bg64 = base64.b64encode(_bgf.read()).decode()
                st.markdown(
                    f"""
                    <style>
                    .stApp {{
                        background-image: url('data:image/png;base64,{_bg64}');
                        background-size: cover;
                        background-position: center;
                        background-repeat: no-repeat;
                        background-attachment: fixed;
                    }}
                    </style>
                    """,
                    unsafe_allow_html=True,
                )
        except Exception:
            pass

        # Left-aligned login with measured width and spacing
        st.markdown(
            """
            <style>
            /* Login layout tweaks */
            .login-container-title { text-align: left; font-size: 1.8rem; font-weight: 800; margin: 6vh 0 0.75rem 4vw; max-width: 720px; }
            .login-section { max-width: 720px; margin-left: 4vw; }
            /* Make inputs visually consistent and wider on large screens */
            .login-section .stTextInput>div>div>input { height: 40px; }
            .login-section .stButton>button { height: 38px; padding: 0 18px; }
            @media (max-width: 900px) {
                .login-container-title { margin-left: 2vw; max-width: 94vw; }
                .login-section { margin-left: 2vw; max-width: 94vw; }
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        left_col, _right_space = st.columns([2.2, 3.8])
        with left_col:
            st.markdown(
                """
                <div class="login-container-title">AI Powered Knowledge Engine for Smart Support and Ticket Resolution</div>
                """,
                unsafe_allow_html=True,
            )
            with st.container():
                st.markdown("<div class='login-section'>", unsafe_allow_html=True)
                auth_mode = st.radio("Authentication Mode", ["Login", "Register"], horizontal=True, index=0, key="auth_mode", label_visibility="collapsed")

                if auth_mode == "Login":
                    st.header("🔐 Login")
                    u = st.text_input("Username", value="", placeholder="Enter username", key="login_user")
                    p = st.text_input("Password", value="", placeholder="Enter password", type="password", key="login_pass")
                    login_clicked = st.button("Login", type="primary")
                else:
                    st.header("🆕 Register")
                    ru = st.text_input("Username", value="", placeholder="Choose a username", key="reg_user")
                    rp = st.text_input("Password", value="", placeholder="Create a password", type="password", key="reg_pass")
                    rc = st.text_input("Confirm Password", value="", placeholder="Re-enter password", type="password", key="reg_confirm")
                    register_clicked = st.button("Create Account", type="primary")
                st.markdown("</div>", unsafe_allow_html=True)

        # Handle auth actions
        if 'auth_mode' in st.session_state and st.session_state.auth_mode == 'Register':
            if 'register_clicked' in locals() and register_clicked:
                if not ru or not rp or not rc:
                    st.error("Please fill all fields.")
                elif rp != rc:
                    st.error("Passwords do not match.")
                else:
                    ok, msg = create_user(ru, rp)
                    if ok:
                        st.success(msg)
                        st.session_state.authenticated = True
                        st.session_state.auth_user = ru
                        st.rerun()
                    else:
                        st.error(msg)
        else:
            if 'login_clicked' in locals() and login_clicked:
                if any_registered_users():
                    if verify_user(u, p):
                        st.session_state.authenticated = True
                        st.session_state.auth_user = u
                        st.success("Logged in.")
                        st.rerun()
                    else:
                        st.error("Invalid credentials.")
                else:
                    if app_user and app_pass:
                        if u == app_user and p == app_pass:
                            st.session_state.authenticated = True
                            st.session_state.auth_user = u
                            st.success("Logged in.")
                            st.rerun()
                        else:
                            st.error("Invalid credentials.")
                    else:
                        if u and p:
                            st.session_state.authenticated = True
                            st.session_state.auth_user = u
                            st.info("No APP_USERNAME/APP_PASSWORD set; allowing temporary access.")
                            st.rerun()
                        else:
                            st.error("Please provide credentials.")
        st.stop()
    
    # Create sidebar
    settings = create_sidebar()
    # Sidebar logout
    with st.sidebar:
        st.markdown("---")
        st.caption(f"Signed in as: {st.session_state.auth_user or 'user'}")
        if st.button("🚪 Logout"):
            st.session_state.authenticated = False
            st.session_state.auth_user = None
            st.experimental_set_query_params()
            st.rerun()
    
    # Display main header
    display_main_header()
    
    # Create tabs (Data Integration removed)
    tab1, tab2, tab3, tab5 = st.tabs([
        "💬 Query Resolution", 
        "🤖 AI Chatbot",
        "🎫 Ticket Management",
        "📈 Analytics"
    ])
    
    with tab1:
        create_query_resolution_tab()
    
    with tab2:
        create_chatbot_tab()
    
    with tab3:
        create_ticket_management_tab()
    
    with tab5:
        create_analytics_tab()

if __name__ == "__main__":
    main()
